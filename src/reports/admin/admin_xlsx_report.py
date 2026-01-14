from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
import pytz


def build_admin_xlsx(stats: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Administrativo"

    # ============================
    # CONFIGURACIÓN DE COLUMNAS
    # ============================
    widths = [30, 18, 18, 18, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ============================
    # HEADER (ZONA SEGURA)
    # ============================
    # Reservamos filas 1 a 8 SOLO para header
    header_end_row = 8

    logo_path = os.path.join(
        os.getcwd(),
        "src",
        "assets",
        "logos",
        "logoFinopsLatam.png"
    )

    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 120
        logo.height = 120
        ws.add_image(logo, "A1")

    # Fecha arriba derecha
    chile_tz = pytz.timezone("America/Santiago")
    now_cl = datetime.now(chile_tz).strftime("%d/%m/%Y %H:%M CLT")

    ws.merge_cells("E1:G1")
    ws["E1"] = f"Generado el {now_cl}"
    ws["E1"].alignment = Alignment(horizontal="right")
    ws["E1"].font = Font(size=10)

    # Título centrado
    ws.merge_cells("C4:F4")
    ws["C4"] = "Reporte Administrativo — FinOpsLatam"
    ws["C4"].font = Font(size=16, bold=True)
    ws["C4"].alignment = Alignment(horizontal="center")

    # ============================
    # KPI TABLE (EMPIEZA MÁS ABAJO)
    # ============================
    table_start = header_end_row + 2  # fila 10

    ws[f"A{table_start}"] = "Métrica"
    ws[f"B{table_start}"] = "Valor"
    ws[f"C{table_start}"] = "Estado"
    ws[f"D{table_start}"] = "Porcentaje"

    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{table_start}"].font = Font(bold=True)

    total = stats["total_users"]
    active = stats["active_users"]
    inactive = stats["inactive_users"]

    active_pct = round((active / total) * 100, 1) if total else 0
    inactive_pct = round((inactive / total) * 100, 1) if total else 0

    rows = [
        ("Usuarios totales", total, "", ""),
        ("Usuarios activos", active, "Activo", f"{active_pct}%"),
        ("Usuarios inactivos", inactive, "Inactivo", f"{inactive_pct}%"),
    ]

    for i, row in enumerate(rows, start=table_start + 1):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"], ws[f"D{i}"] = row

    # ============================
    # PIE CHART (DEBAJO KPI)
    # ============================
    pie = PieChart()
    pie.title = "Usuarios activos vs inactivos"

    labels = Reference(ws, min_col=1, min_row=table_start + 2, max_row=table_start + 3)
    values = Reference(ws, min_col=2, min_row=table_start + 2, max_row=table_start + 3)

    pie.add_data(values, titles_from_data=False)
    pie.set_categories(labels)

    pie_anchor_row = table_start + 6
    ws.add_chart(pie, f"F{pie_anchor_row}")

    # ============================
    # PLAN TABLE (MUCHO MÁS ABAJO)
    # ============================
    plan_table_start = pie_anchor_row + 18  # separación REAL

    ws[f"A{plan_table_start}"] = "Plan"
    ws[f"B{plan_table_start}"] = "Cantidad"
    ws[f"A{plan_table_start}"].font = ws[f"B{plan_table_start}"].font = Font(bold=True)

    for i, p in enumerate(stats["users_by_plan"], start=plan_table_start + 1):
        ws[f"A{i}"] = p["plan"]
        ws[f"B{i}"] = p["count"]

    # ============================
    # BAR CHART (DEBAJO TABLA PLAN)
    # ============================
    bar = BarChart()
    bar.title = "Usuarios por plan"
    bar.y_axis.title = "Cantidad"
    bar.x_axis.title = "Plan"

    data_ref = Reference(
        ws,
        min_col=2,
        min_row=plan_table_start,
        max_row=plan_table_start + len(stats["users_by_plan"]),
    )

    cats_ref = Reference(
        ws,
        min_col=1,
        min_row=plan_table_start + 1,
        max_row=plan_table_start + len(stats["users_by_plan"]),
    )

    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)

    ws.add_chart(bar, f"F{plan_table_start + 4}")

    # ============================
    # FOOTER
    # ============================
    footer_row = plan_table_start + len(stats["users_by_plan"]) + 20
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws[f"A{footer_row}"] = (
        "© 2026 FinOpsLatam — Información confidencial. Uso exclusivo interno."
    )
    ws[f"A{footer_row}"].alignment = Alignment(horizontal="center")
    ws[f"A{footer_row}"].font = Font(size=9)

    # ============================
    # EXPORT
    # ============================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.read()
