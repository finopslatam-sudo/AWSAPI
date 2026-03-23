"""
COST XLSX REPORT
================
3 hojas: Resumen KPIs / Tendencia Mensual / Distribución por Servicio
"""

from io import BytesIO
from datetime import datetime
import pytz

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.services.client_dashboard_service import ClientDashboardService
from src.services.client_stats_service import get_users_by_client, get_client_plan
from src.models.aws_account import AWSAccount


def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_style():
    return Font(color="FFFFFF", bold=True, size=9)


def _apply_table_header(ws, row, cols, fill_color="1E293B"):
    fill = PatternFill("solid", fgColor=fill_color)
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _hdr_style()
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()


def _apply_data_row(ws, row, values, alt=False):
    fill = PatternFill("solid", fgColor="F8FAFC") if alt else PatternFill("solid", fgColor="FFFFFF")
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(size=9, color="0F172A")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _border()


def build_cost_xlsx(client_id: int, aws_account_id: int | None = None) -> bytes:

    cost     = ClientDashboardService.get_cost_data(client_id, aws_account_id)
    plan     = get_client_plan(client_id) or "Sin plan"
    users    = get_users_by_client(client_id)
    acc_count = AWSAccount.query.filter_by(client_id=client_id, is_active=True).count()

    account_label = "Todas las cuentas"
    if aws_account_id:
        acc = AWSAccount.query.get(aws_account_id)
        account_label = acc.account_name if acc else str(aws_account_id)

    prev_month   = float(cost.get("previous_month_cost", 0))
    curr_partial = float(cost.get("current_month_partial", 0))
    prev_year    = float(cost.get("previous_year_cost", 0))
    curr_year    = float(cost.get("current_year_ytd", 0))
    savings      = float(cost.get("potential_savings", 0))
    ann_savings  = float(cost.get("annual_estimated_savings", 0))
    sav_pct      = float(cost.get("savings_percentage", 0))
    monthly      = cost.get("monthly_cost", [])
    svc          = sorted(cost.get("service_breakdown", []),
                          key=lambda x: float(x.get("amount", 0)), reverse=True)

    chile_tz = pytz.timezone("America/Santiago")
    generated = datetime.now(chile_tz).strftime("%d/%m/%Y %H:%M CLT")

    dark_fill    = PatternFill("solid", fgColor="1E293B")
    summary_fill = PatternFill("solid", fgColor="EFF6FF")
    green_fill   = PatternFill("solid", fgColor="F0FDF4")
    label_font   = Font(bold=True, color="0F172A", size=9)
    value_font   = Font(color="334155", size=9)
    title_font   = Font(bold=True, size=14, color="0F172A")
    sub_font     = Font(size=9, color="64748B")
    thin         = _border()

    wb = Workbook()

    # ═══════════════════════════════════════
    # HOJA 1 — RESUMEN FINANCIERO
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Financiero"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 22

    # Título
    ws1["A1"] = "Reporte de Costos — FinOpsLatam"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:B1")
    ws1["A2"] = f"Generado: {generated}"
    ws1["A2"].font = sub_font
    ws1.merge_cells("A2:B2")

    ws1["A4"] = "INFORMACIÓN GENERAL"
    ws1["A4"].font = Font(bold=True, size=10, color="1E293B")

    meta = [
        ("Plan de suscripción", plan),
        ("Cuentas AWS activas", f"{acc_count}  ({account_label})"),
        ("Usuarios activos",    str(users)),
    ]
    for i, (lbl, val) in enumerate(meta, 5):
        ws1.cell(row=i, column=1, value=lbl).font = label_font
        ws1.cell(row=i, column=1).fill = summary_fill
        ws1.cell(row=i, column=1).border = thin
        ws1.cell(row=i, column=2, value=val).font = value_font
        ws1.cell(row=i, column=2).fill = summary_fill
        ws1.cell(row=i, column=2).border = thin

    ws1["A9"] = "KPIs FINANCIEROS"
    ws1["A9"].font = Font(bold=True, size=10, color="1E293B")

    kpis = [
        ("Gasto Mes Anterior",      f"USD ${prev_month:.2f}",  "Mes cerrado"),
        ("Gasto Mes Actual (YTD)",  f"USD ${curr_partial:.2f}", "Parcial mes en curso"),
        ("Ahorro Mensual Acumulado", f"USD ${savings:.2f}",    f"{sav_pct:.1f}% del gasto"),
        ("Gasto Año Anterior",      f"USD ${prev_year:.2f}",   "Año fiscal anterior"),
        ("Gasto Año Actual (YTD)",  f"USD ${curr_year:.2f}",   "Acumulado año en curso"),
        ("Ahorro Anual Estimado",   f"USD ${ann_savings:.2f}", "Savings × 12 meses"),
    ]
    for i, (lbl, val, note) in enumerate(kpis, 10):
        ws1.cell(row=i, column=1, value=lbl).font = label_font
        ws1.cell(row=i, column=1).fill = green_fill
        ws1.cell(row=i, column=1).border = thin
        cell_v = ws1.cell(row=i, column=2, value=val)
        cell_v.font = Font(bold=True, size=10, color="15803D")
        cell_v.fill = green_fill
        cell_v.border = thin
        ws1.cell(row=i, column=3, value=note).font = sub_font

    # ═══════════════════════════════════════
    # HOJA 2 — TENDENCIA MENSUAL
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("Tendencia Mensual")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    ws2["A1"] = "Tendencia de Costos — Últimos 6 Meses"
    ws2["A1"].font = Font(bold=True, size=12, color="0F172A")
    ws2.merge_cells("A1:D1")

    _apply_table_header(ws2, 3, ["Mes", "Gasto (USD)", "% del Total", "Variación"])
    total_mc = sum(float(m["amount"]) for m in monthly) or 1
    prev_amt = None
    for r, m in enumerate(monthly[-6:], 4):
        amt = float(m["amount"])
        pct = amt / total_mc * 100
        if prev_amt is not None and prev_amt > 0:
            var = f"+{((amt - prev_amt) / prev_amt * 100):.1f}%" if amt >= prev_amt else f"{((amt - prev_amt) / prev_amt * 100):.1f}%"
        else:
            var = "—"
        _apply_data_row(ws2, r, [m["month"], round(amt, 2), f"{pct:.1f}%", var], alt=(r % 2 == 0))
        prev_amt = amt

    # ═══════════════════════════════════════
    # HOJA 3 — DISTRIBUCIÓN POR SERVICIO
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("Por Servicio")
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 14

    ws3["A1"] = "Distribución de Costos por Servicio (Mes Actual)"
    ws3["A1"].font = Font(bold=True, size=12, color="0F172A")
    ws3.merge_cells("A1:C1")

    _apply_table_header(ws3, 3, ["Servicio", "Costo (USD)", "% del Total"])
    total_svc = sum(float(s.get("amount", 0)) for s in svc) or 1
    for r, s in enumerate(svc, 4):
        amt = float(s.get("amount", 0))
        pct = amt / total_svc * 100
        _apply_data_row(ws3, r, [s.get("service", ""), round(amt, 2), f"{pct:.1f}%"], alt=(r % 2 == 0))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
