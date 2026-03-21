"""
CLIENT XLSX REPORT
==================
Hoja única: resumen arriba + tabla de findings completa abajo.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import pytz


RESOLUTION = {
    "EBS_GP2_TO_GP3":                  "Migrar volumen de GP2 a GP3 via modify-volume. Sin downtime.",
    "EC2_UNDERUTILIZED":               "Rightsizing: reducir tipo de instancia o apagarla si no esta en uso.",
    "STOPPED_INSTANCE":                "Crear AMI de respaldo y terminar la instancia para detener el cobro.",
    "UNATTACHED_VOLUME":               "Tomar snapshot y eliminar el volumen desconectado.",
    "CLOUDWATCH_NO_RETENTION":         "Configurar politica de retencion en el Log Group (ej. 30-90 dias).",
    "CLOUDWATCH_HIGH_RETENTION":       "Reducir retencion del Log Group o mover logs frios a S3 Glacier.",
    "CLOUDWATCH_STORAGE_RIGHTSIZING":  "Reducir retencion o filtrar metricas enviadas a CloudWatch.",
    "LAMBDA_HIGH_MEMORY":              "Reducir memoria de la funcion Lambda con AWS Lambda Power Tuning.",
    "LAMBDA_MEMORY_RIGHTSIZING":       "Ajustar memoria de Lambda al valor optimo con Power Tuning.",
    "LAMBDA_DEPRECATED_RUNTIME":       "Actualizar el runtime de Lambda a una version soportada.",
    "RDS_UNDERUTILIZED":               "Escalar hacia abajo la instancia RDS o usar Aurora Serverless.",
    "RDS_GP2_STORAGE":                 "Migrar almacenamiento RDS de GP2 a GP3 para reducir costo.",
    "RDS_MULTI_AZ_DISABLED":           "Habilitar Multi-AZ para alta disponibilidad (solo produccion).",
    "RDS_NOT_ENCRYPTED":               "Restaurar desde snapshot con cifrado habilitado.",
    "RDS_NO_BACKUP_RETENTION":         "Configurar retencion de backups automaticos a minimo 7 dias.",
    "RDS_PUBLIC_ACCESS":               "Desactivar acceso publico y usar VPC o bastion host.",
    "DYNAMODB_EMPTY_TABLE":            "Exportar a S3 si hay datos y eliminar la tabla vacia.",
    "DYNAMODB_PROVISIONED_MODE":       "Cambiar a On-Demand o ajustar las unidades de capacidad.",
    "DYNAMODB_PROVISIONED_RIGHTSIZING":"Ajustar Read/Write Capacity Units al consumo real.",
    "NAT_IDLE_GATEWAY":                "Eliminar NAT Gateway sin trafico (verificar rutas antes).",
    "S3_STORAGE_RIGHTSIZING_REVIEW":   "Aplicar Intelligent-Tiering o ciclo de vida para objetos frios.",
    "EC2_RI":                          "Adquirir Reserved Instances de 1 anio para uso sostenido (-40%).",
    "RI_UNUSED":                       "Vender RIs no usadas en AWS Marketplace o reasignarlas.",
    "LOW_RI_COVERAGE":                 "Aumentar cobertura de Reserved Instances en instancias de uso alto.",
    "SP_REVIEW":                       "Evaluar Savings Plans para workloads con uso predecible.",
    "ECS_SERVICE_RIGHTSIZING_REVIEW":  "Revisar limites de CPU/memoria en tareas ECS.",
    "EKS_NODEGROUP_RIGHTSIZING_REVIEW":"Ajustar tipo de instancia y tamanio del Node Group de EKS.",
    "REDSHIFT_UNDERUTILIZED":          "Pausar el cluster o migrar a Redshift Serverless.",
    "RIGHTSIZING_OPPORTUNITY":         "Migrar a tipo/tamanio inferior que cubra la carga real.",
}

def _resolution(finding_type: str) -> str:
    return RESOLUTION.get(
        finding_type,
        "Revisar el recurso en la consola AWS y evaluar optimizacion."
    )

def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def build_client_xlsx(stats: dict) -> bytes:

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings & Rightsizing"

    chile_tz    = pytz.timezone("America/Santiago")
    generated_at = datetime.now(chile_tz).strftime("%d/%m/%Y %H:%M CLT")

    thin        = _border()
    dark_fill   = PatternFill("solid", fgColor="1E293B")
    alt_fill    = PatternFill("solid", fgColor="F8FAFC")
    summary_fill= PatternFill("solid", fgColor="EFF6FF")
    hdr_font    = Font(color="FFFFFF", bold=True, size=9)
    label_font  = Font(bold=True, color="0F172A", size=9)
    value_font  = Font(color="334155", size=9)
    title_font  = Font(bold=True, size=14, color="0F172A")
    sub_font    = Font(size=9, color="64748B")

    # ── Anchos de columna (11 columnas de findings) ──────────────────────
    col_configs = [
        ("Account",       26),
        ("Service",       14),
        ("Type",          26),
        ("Resource",      36),
        ("Region",        13),
        ("Savings (USD)", 14),
        ("Status",        11),
        ("Severity",      11),
        ("Finding",       48),
        ("How to Fix",    52),
        ("Detected At",   14),
    ]
    for idx, (_, w) in enumerate(col_configs, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ================================================
    # BLOQUE 1 — TÍTULO
    # ================================================
    ws.merge_cells("A1:C1")
    ws["A1"] = "Reporte FinOpsLatam — Findings & Rightsizing"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generado el {generated_at}"
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(horizontal="left")

    # ================================================
    # BLOQUE 2 — RESUMEN (fila 4 en adelante)
    # ================================================
    summary = stats.get("findings_summary") or {}
    savings_val = summary.get("estimated_monthly_savings") or summary.get("savings") or 0

    summary_rows = [
        ("Plan contratado",          stats.get("plan") or "Sin plan activo"),
        ("Cuentas AWS escaneadas",   stats.get("account_count", 0)),
        ("Usuarios asociados",       stats.get("user_count", 0)),
        ("Findings activos",         summary.get("active", 0)),
        ("Findings resueltos",       summary.get("resolved", 0)),
        ("Findings severidad HIGH",  summary.get("high", 0)),
        ("Ahorro mensual estimado",  f"USD ${float(savings_val):.2f}"),
    ]

    # Encabezado resumen
    ws["A4"] = "Métrica"
    ws["B4"] = "Valor"
    ws["A4"].font = hdr_font
    ws["A4"].fill = dark_fill
    ws["A4"].border = thin
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B4"].font = hdr_font
    ws["B4"].fill = dark_fill
    ws["B4"].border = thin
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center")

    for i, (metric, value) in enumerate(summary_rows, start=5):
        ws[f"A{i}"] = metric
        ws[f"A{i}"].font = label_font
        ws[f"A{i}"].border = thin
        ws[f"A{i}"].fill = summary_fill
        ws[f"A{i}"].alignment = Alignment(vertical="center")

        ws[f"B{i}"] = value
        ws[f"B{i}"].font = value_font
        ws[f"B{i}"].border = thin
        ws[f"B{i}"].alignment = Alignment(vertical="center")

    # ================================================
    # BLOQUE 3 — TABLA FINDINGS (debajo del resumen)
    # ================================================
    findings_header_row = 5 + len(summary_rows) + 2   # fila 14

    # Título sección
    ws.merge_cells(f"A{findings_header_row - 1}:C{findings_header_row - 1}")
    ws[f"A{findings_header_row - 1}"] = "Detalle de Findings & Rightsizing"
    ws[f"A{findings_header_row - 1}"].font = Font(bold=True, size=11, color="1E293B")
    ws[f"A{findings_header_row - 1}"].alignment = Alignment(horizontal="left")

    # Encabezados de tabla
    for col_idx, (col_name, _) in enumerate(col_configs, start=1):
        cell = ws.cell(row=findings_header_row, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = dark_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[findings_header_row].height = 18

    # Freeze header de findings para scroll cómodo
    ws.freeze_panes = ws.cell(row=findings_header_row + 1, column=1)

    # Filas de datos
    findings = stats.get("findings") or []
    for r_offset, f in enumerate(findings):
        row = findings_header_row + 1 + r_offset
        savings = float(f.get("estimated_monthly_savings") or 0)
        detected = (f.get("created_at") or f.get("detected_at") or "")[:10]

        row_data = [
            f.get("aws_account_name") or f.get("aws_account_number") or "",
            f.get("aws_service", ""),
            f.get("finding_type", ""),
            f.get("resource_id", ""),
            f.get("region", "") or "",
            f"${savings:.2f}",
            "Resolved" if f.get("resolved") else "Active",
            f.get("severity", ""),
            f.get("message", "") or "",
            _resolution(f.get("finding_type", "")),
            detected,
        ]

        fill = alt_fill if r_offset % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = value_font
            cell.border = thin
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_idx in (9, 10)),
            )
            if fill:
                cell.fill = fill

        ws.row_dimensions[row].height = 32

    # ================================================
    # EXPORT
    # ================================================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
