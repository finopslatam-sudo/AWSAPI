"""
RISK & COMPLIANCE XLSX REPORT
==============================
3 hojas: Resumen Riesgo / Findings Alta Severidad / Riesgo por Servicio
"""

from io import BytesIO
from datetime import datetime
import pytz

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.services.client_stats_service import get_users_by_client, get_client_plan
from src.services.client_findings_service import ClientFindingsService
from src.services.dashboard.governance_service import GovernanceService
from src.services.dashboard.risk_service import RiskService
from src.models.aws_account import AWSAccount


def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(ws, row, cols, fill_color="1E293B"):
    fill = PatternFill("solid", fgColor=fill_color)
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()


def _apply_row(ws, row, values, alt=False, font_colors=None):
    fill = PatternFill("solid", fgColor="FFF5F5" if alt else "FFFFFF")
    thin = _border()
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        color = (font_colors[c - 1] if font_colors and c - 1 < len(font_colors) else "0F172A")
        cell.font = Font(size=9, color=color)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin


def build_risk_xlsx(client_id: int, aws_account_id: int | None = None) -> bytes:

    plan  = get_client_plan(client_id) or "Sin plan"
    users = get_users_by_client(client_id)
    acc_count = AWSAccount.query.filter_by(client_id=client_id, is_active=True).count()

    account_label = "Todas las cuentas"
    if aws_account_id:
        acc = AWSAccount.query.get(aws_account_id)
        account_label = acc.account_name if acc else str(aws_account_id)

    try:
        gov      = GovernanceService.get_governance_score(client_id, aws_account_id)
        risk     = RiskService.get_risk_profile(client_id, aws_account_id)
        risk_bkdn = RiskService.get_risk_breakdown_by_service(client_id, aws_account_id)
    except Exception:
        gov      = {"compliance_percentage": 0}
        risk     = {"risk_level": "N/A", "risk_score": 0, "high": 0, "medium": 0, "low": 0}
        risk_bkdn = []

    stats = ClientFindingsService.get_stats(client_id, aws_account_id)
    high_findings = ClientFindingsService.list_findings(
        client_id=client_id,
        page=1, per_page=50,
        status="active", severity="HIGH",
        finding_type=None, service=None, region=None, search=None,
        sort_by="estimated_monthly_savings", sort_order="desc",
    ).get("data", [])

    compliance = float(gov.get("compliance_percentage", 0))
    risk_level = str(risk.get("risk_level", "N/A")).upper()
    risk_score = float(risk.get("risk_score", 0))

    chile_tz  = pytz.timezone("America/Santiago")
    generated = datetime.now(chile_tz).strftime("%d/%m/%Y %H:%M CLT")

    thin       = _border()
    dark_fill  = PatternFill("solid", fgColor="1E293B")
    red_fill   = PatternFill("solid", fgColor="7F1D1D")
    kpi_fill   = PatternFill("solid", fgColor="FEF2F2")
    sum_fill   = PatternFill("solid", fgColor="FFF5F5")
    label_font = Font(bold=True, color="0F172A", size=9)
    value_font = Font(color="334155", size=9)
    title_font = Font(bold=True, size=14, color="0F172A")
    sub_font   = Font(size=9, color="64748B")

    wb = Workbook()

    # ═══════════════════════════════════════
    # HOJA 1 — RESUMEN DE RIESGO
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen de Riesgo"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 22

    ws1["A1"] = "Reporte de Riesgo & Compliance — FinOpsLatam"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:B1")
    ws1["A2"] = f"Generado: {generated}"
    ws1["A2"].font = sub_font
    ws1.merge_cells("A2:B2")

    ws1["A4"] = "INFORMACIÓN GENERAL"
    ws1["A4"].font = Font(bold=True, size=10, color="1E293B")

    for i, (lbl, val) in enumerate([
        ("Plan de suscripción",  plan),
        ("Cuentas AWS activas",  f"{acc_count}  ({account_label})"),
        ("Usuarios activos",     str(users)),
    ], 5):
        ws1.cell(row=i, column=1, value=lbl).font = label_font
        ws1.cell(row=i, column=1).fill = sum_fill
        ws1.cell(row=i, column=1).border = thin
        ws1.cell(row=i, column=2, value=val).font = value_font
        ws1.cell(row=i, column=2).fill = sum_fill
        ws1.cell(row=i, column=2).border = thin

    ws1["A9"] = "KPIs DE RIESGO & COMPLIANCE"
    ws1["A9"].font = Font(bold=True, size=10, color="1E293B")

    risk_color = {
        "LOW": "16A34A", "MEDIUM": "D97706",
        "HIGH": "DC2626", "CRITICAL": "9F1239",
    }.get(risk_level, "334155")

    kpis = [
        ("Nivel de Riesgo",            risk_level,                "LOW / MEDIUM / HIGH / CRITICAL"),
        ("Score de Riesgo",            f"{risk_score:.0f} / 100", "Mayor score = menor riesgo"),
        ("Cumplimiento (Compliance)",  f"{compliance:.1f}%",      "Óptimo ≥80% · Atención ≥50% · Crítico <50%"),
        ("Findings Alta Severidad",    str(risk.get("high", 0)),  "Requieren acción inmediata"),
        ("Findings Media Severidad",   str(risk.get("medium", 0)),"Atención en el corto plazo"),
        ("Findings Baja Severidad",    str(risk.get("low", 0)),   "Monitorear"),
        ("Total Findings",             str(stats.get("total", 0)), "Detectados en el entorno"),
        ("Findings Activos",           str(stats.get("active", 0)), "Pendientes de resolución"),
        ("Findings Resueltos",         str(stats.get("resolved", 0)), "Optimizaciones aplicadas"),
    ]
    for i, (lbl, val, note) in enumerate(kpis, 10):
        ws1.cell(row=i, column=1, value=lbl).font = label_font
        ws1.cell(row=i, column=1).fill = kpi_fill
        ws1.cell(row=i, column=1).border = thin
        fg = risk_color if lbl == "Nivel de Riesgo" else "DC2626" if "Alta" in lbl else "0F172A"
        cell_v = ws1.cell(row=i, column=2, value=val)
        cell_v.font = Font(bold=True, size=10, color=fg)
        cell_v.fill = kpi_fill
        cell_v.border = thin
        ws1.cell(row=i, column=3, value=note).font = sub_font

    # ═══════════════════════════════════════
    # HOJA 2 — FINDINGS ALTA SEVERIDAD
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("Findings Alta Severidad")
    for col, w in zip("ABCDEF", [20, 35, 25, 16, 50, 40]):
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[1].height = 20

    ws2["A1"] = "Findings de Alta Severidad — Requieren Acción Inmediata"
    ws2["A1"].font = Font(bold=True, size=12, color="7F1D1D")
    ws2.merge_cells("A1:F1")

    _apply_header(ws2, 3,
                  ["Servicio", "Recurso", "Tipo", "Ahorro/mes (USD)", "Descripción", "Cuenta AWS"],
                  fill_color="7F1D1D")

    for r, f in enumerate(high_findings, 4):
        colors_per_col = ["0F172A", "64748B", "64748B", "DC2626", "0F172A", "64748B"]
        _apply_row(ws2, r, [
            f.get("aws_service", ""),
            f.get("resource_id", ""),
            f.get("finding_type", ""),
            round(float(f.get("estimated_monthly_savings", 0)), 2),
            f.get("message", ""),
            f.get("aws_account_name", "") or f.get("aws_account_number", ""),
        ], alt=(r % 2 == 0), font_colors=colors_per_col)
    ws2.freeze_panes = "A4"

    # ═══════════════════════════════════════
    # HOJA 3 — RIESGO POR SERVICIO
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("Riesgo por Servicio")
    for col, w in zip("ABCDEF", [35, 18, 14, 14, 14, 14]):
        ws3.column_dimensions[col].width = w

    ws3["A1"] = "Distribución de Riesgo por Servicio AWS"
    ws3["A1"].font = Font(bold=True, size=12, color="0F172A")
    ws3.merge_cells("A1:F1")

    _apply_header(ws3, 3,
                  ["Servicio", "Total Recursos", "Alta Severidad", "Media", "Baja", "Score"])

    for r, s in enumerate(risk_bkdn[:30], 4):
        total_r = int(s.get("total_resources", 0))
        h = int(s.get("high", 0))
        m = int(s.get("medium", 0))
        l = int(s.get("low", 0))
        points = h * 5 + m * 3 + l
        score = round(100 - (points / max(total_r * 5, 1) * 100), 1)
        _apply_row(ws3, r, [
            s.get("service_name", ""), total_r, h, m, l, score
        ], alt=(r % 2 == 0),
        font_colors=["0F172A", "0F172A", "DC2626", "D97706", "16A34A", "1D4ED8"])
    ws3.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
