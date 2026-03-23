"""
Helper functions for the inventory XLSX report.
Column-width setter, header/data row writers, KPI block, and fill selectors.
"""

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .styles import (
    ALT_FILL, WHITE_FILL, DARK_FILL, THIN,
    HDR_FONT, VALUE_FONT,
    _fill,
)


# ─────────────────────────────────────────────────────────
# LAYOUT
# ─────────────────────────────────────────────────────────

def _set_col_widths(ws, widths: list[tuple[int, float]]):
    for col_idx, w in widths:
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ─────────────────────────────────────────────────────────
# ROW WRITERS
# ─────────────────────────────────────────────────────────

def _write_header_row(ws, row: int, labels: list[str], fill=None):
    f = fill or DARK_FILL
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HDR_FONT
        cell.fill = f
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20


def _write_data_row(ws, row: int, values: list, alt: bool = False,
                    wrap_cols: set | None = None, row_h: int = 18,
                    color_map: dict | None = None):
    bg = ALT_FILL if alt else WHITE_FILL
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = VALUE_FONT
        cell.fill = color_map.get(c, bg) if color_map else bg
        cell.border = THIN
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(wrap_cols is not None and c in wrap_cols),
        )
    ws.row_dimensions[row].height = row_h


# ─────────────────────────────────────────────────────────
# KPI BLOCK
# ─────────────────────────────────────────────────────────

def _kpi_block(ws, start_row: int, start_col: int, label: str, value: str,
               note: str, fill: PatternFill, val_color: str = "1D4ED8"):
    """Escribe un bloque KPI de 3 filas × 1 columna."""
    r = start_row
    c = start_col

    lbl_cell  = ws.cell(row=r,   column=c, value=label)
    val_cell  = ws.cell(row=r+1, column=c, value=value)
    note_cell = ws.cell(row=r+2, column=c, value=note)

    lbl_cell.font  = Font(size=8, color="64748B")
    val_cell.font  = Font(bold=True, size=14, color=val_color)
    note_cell.font = Font(size=7, color="94A3B8")

    for cell in (lbl_cell, val_cell, note_cell):
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[r].height   = 14
    ws.row_dimensions[r+1].height = 22
    ws.row_dimensions[r+2].height = 13


# ─────────────────────────────────────────────────────────
# CONDITIONAL FILLS
# ─────────────────────────────────────────────────────────

def _severity_fill(severity: str) -> PatternFill:
    s = (severity or "").upper()
    if s == "CRITICAL":
        return _fill("FFE4E6")
    if s == "HIGH":
        return _fill("FEE2E2")
    if s == "MEDIUM":
        return _fill("FEF3C7")
    if s == "LOW":
        return _fill("DCFCE7")
    return WHITE_FILL


def _state_fill(state: str) -> PatternFill:
    s = (state or "").lower()
    if s in ("running", "available", "active"):
        return _fill("DCFCE7")
    if s in ("stopped", "stopping"):
        return _fill("FEE2E2")
    if s in ("idle", "inactive"):
        return _fill("FEF3C7")
    return ALT_FILL
