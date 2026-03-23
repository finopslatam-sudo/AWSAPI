"""Hoja por servicio AWS."""

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment

from .styles import TITLE_FONT, SUB_FONT, SUMMARY_FILL, LABEL_FONT
from .helpers import _set_col_widths, _write_header_row, _write_data_row


def build_service_sheet(wb, *, generated: str, total: int,
                        by_service: dict, resources: list):
    ws = wb.create_sheet("Por Servicio AWS")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [(1, 32), (2, 14), (3, 14), (4, 28)])

    ws.merge_cells("A1:D1")
    ws["A1"] = "Inventario por Servicio AWS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generado: {generated}"
    ws["A2"].font = SUB_FONT

    _write_header_row(ws, 4, ["Servicio AWS", "Recursos", "% del Total", "Distribución"])

    total_s = max(total, 1)
    svc_chart_data = []
    for i, (svc, count) in enumerate(by_service.items()):
        row = 5 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 3)) + "░" * (33 - max(1, int(pct / 3)))
        _write_data_row(ws, row,
                        [svc, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))
        svc_chart_data.append((svc, count))

    total_row = 5 + len(by_service)
    _write_data_row(ws, total_row,
                    ["TOTAL", total, "100%", ""],
                    color_map={1: SUMMARY_FILL, 2: SUMMARY_FILL, 3: SUMMARY_FILL, 4: SUMMARY_FILL})
    for c in range(1, 5):
        ws.cell(total_row, c).font = LABEL_FONT

    if svc_chart_data:
        chart = BarChart()
        chart.type     = "bar"
        chart.title    = "Recursos por Servicio AWS"
        chart.y_axis.title = "Servicio"
        chart.x_axis.title = "Recursos"
        chart.style    = 10
        chart.width    = 18
        chart.height   = max(8, len(svc_chart_data) * 0.65)
        chart.grouping = "clustered"

        data_ref = Reference(ws, min_col=2, min_row=4,
                             max_col=2, max_row=4 + len(svc_chart_data))
        cats_ref = Reference(ws, min_col=1, min_row=5,
                             max_row=4 + len(svc_chart_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "2563EB"
        ws.add_chart(chart, "F4")
