"""Resumen ejecutivo para el XLSX de inventario."""

from openpyxl.styles import Alignment

from .styles import (
    TITLE_FONT, SUB_FONT, SECTION_FONT, LABEL_FONT,
    ACCENT_FILL, BLUE_FILL, RED_FILL, GREEN_FILL,
    AMBER_FILL, PURPLE_FILL, SUMMARY_FILL, THIN, _fill,
)
from .helpers import (
    _set_col_widths, _write_header_row, _write_data_row,
    _kpi_block, _severity_fill, _state_fill,
)


def build_summary_sheet(ws, *, generated: str, plan: str, acc_count: int,
                        acc_label: str, users: int, total: int, with_f: int,
                        without_f: int, active_f_count: int, total_savings: float,
                        by_service: dict, by_state: dict, by_region: dict):

    ws.title = "Resumen Ejecutivo"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [
        (1, 4), (2, 28), (3, 22), (4, 4), (5, 28), (6, 22), (7, 4), (8, 28), (9, 22),
    ])

    # — título principal —
    ws.merge_cells("B1:I1")
    ws["B1"] = "Inventario de Recursos AWS — FinOpsLatam"
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B2:I2")
    ws["B2"] = (
        f"Generado: {generated}  ·  Plan: {plan}  ·  "
        f"Cuentas: {acc_count} ({acc_label})  ·  Usuarios: {users}"
    )
    ws["B2"].font = SUB_FONT
    ws["B2"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 16

    # — separador —
    for col in range(2, 10):
        ws.cell(row=3, column=col).fill = _fill("2563EB")
    ws.row_dimensions[3].height = 3

    # — bloque KPIs —
    total_s = max(total, 1)
    kpi_data = [
        (5, 2, "TOTAL RECURSOS",      str(total),           "activos en inventario",                               BLUE_FILL,          "1D4ED8"),
        (5, 5, "CON HALLAZGOS",       str(with_f),          f"{(with_f/total_s*100):.1f}% del inventario",         RED_FILL,           "DC2626"),
        (5, 8, "SIN HALLAZGOS",       str(without_f),       f"{(without_f/total_s*100):.1f}% sin alertas",         GREEN_FILL,         "15803D"),
        (9, 2, "HALLAZGOS ACTIVOS",   str(active_f_count),  "findings sin resolver",                               AMBER_FILL,         "B45309"),
        (9, 5, "AHORRO POTENCIAL",    f"USD ${total_savings:,.2f}", "estimado mensual",                            _fill("F0FDF4"),    "059669"),
        (9, 8, "SERVICIOS DISTINTOS", str(len(by_service)), "tipos de servicio AWS",                               PURPLE_FILL,        "6D28D9"),
    ]
    for r, c, lbl, val, note, fill, color in kpi_data:
        for dr in range(3):
            ws.merge_cells(
                start_row=r+dr, start_column=c,
                end_row=r+dr,   end_column=c+1,
            )
        _kpi_block(ws, r, c, lbl, val, note, fill, color)
        for dr in range(3):
            for dc in range(2):
                ws.cell(row=r+dr, column=c+dc).border = THIN

    # — distribución por estado —
    state_start = 13
    ws.merge_cells(f"B{state_start}:I{state_start}")
    ws[f"B{state_start}"] = "Distribución por Estado"
    ws[f"B{state_start}"].font = SECTION_FONT
    ws.row_dimensions[state_start].height = 20

    _write_header_row(ws, state_start + 1, ["Estado", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    for i, (state, count) in enumerate(by_state.items()):
        r = state_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws, r,
                        [state.capitalize(), count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0),
                        color_map={1: _state_fill(state)})

    # — distribución por servicio (top 12) —
    svc_start = state_start + 2 + len(by_state) + 2
    ws.merge_cells(f"B{svc_start}:I{svc_start}")
    ws[f"B{svc_start}"] = "Top Servicios AWS por Recursos"
    ws[f"B{svc_start}"].font = SECTION_FONT
    ws.row_dimensions[svc_start].height = 20

    _write_header_row(ws, svc_start + 1,
                      ["Servicio AWS", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    top_svcs = list(by_service.items())[:12]
    for i, (svc, count) in enumerate(top_svcs):
        r = svc_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws, r,
                        [svc, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))

    # — distribución por región (top 12) —
    reg_start = svc_start + 2 + len(top_svcs) + 2
    ws.merge_cells(f"B{reg_start}:I{reg_start}")
    ws[f"B{reg_start}"] = "Distribución por Región Geográfica"
    ws[f"B{reg_start}"].font = SECTION_FONT
    ws.row_dimensions[reg_start].height = 20

    _write_header_row(ws, reg_start + 1,
                      ["Región AWS", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    top_regs = list(by_region.items())[:12]
    for i, (reg, count) in enumerate(top_regs):
        r = reg_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws, r,
                        [reg, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))
