"""Hoja de recursos con hallazgos."""

from openpyxl.styles import Alignment

from .styles import TITLE_FONT, SUB_FONT, _fill
from .helpers import _set_col_widths, _write_header_row, _write_data_row, _severity_fill


def build_findings_sheet(wb, *, generated: str, total_savings: float, resources: list):
    ws = wb.create_sheet("Con Hallazgos")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [
        (1, 22), (2, 15), (3, 20), (4, 36),
        (5, 16), (6, 13), (7, 11), (8, 16),
        (9, 22), (10, 18), (11, 48),
    ])

    ws.merge_cells("A1:K1")
    ws["A1"] = "Recursos con Hallazgos Activos"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    flagged = [r for r in resources if r.get("has_findings")]

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        f"Generado: {generated}  ·  "
        f"{len(flagged)} recursos con hallazgos activos  ·  "
        f"Ahorro potencial total: USD ${total_savings:,.2f}/mes"
    )
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[2].height = 14

    cols = [
        "Cuenta AWS", "Servicio", "Tipo", "ID de Recurso",
        "Región", "Estado", "Nº Hallazgos", "Ahorro Est. (USD/mes)",
        "Tipo de Finding", "Severidad", "Descripción",
    ]
    _write_header_row(ws, 4, cols, fill=_fill("DC2626"))
    ws.freeze_panes = "A5"

    data_row = 5
    for r in flagged:
        findings = r.get("findings", [])
        if not findings:
            continue
        for j, f in enumerate(findings):
            sev = f.get("severity", "—").upper()
            _write_data_row(ws, data_row, [
                r.get("account_name", "")           if j == 0 else "",
                r.get("service_name", "")           if j == 0 else "",
                r.get("resource_type", "")          if j == 0 else "",
                r.get("resource_id", "")            if j == 0 else "",
                r.get("region", "")                 if j == 0 else "",
                r.get("state", "").capitalize()     if j == 0 else "",
                r.get("findings_count", 0)           if j == 0 else "",
                f"${r.get('est_savings', 0):,.2f}"  if j == 0 else "",
                f.get("type", ""),
                sev,
                f.get("message", ""),
            ], alt=(data_row % 2 == 0),
               wrap_cols={11},
               row_h=22,
               color_map={10: _severity_fill(sev)})
            data_row += 1

