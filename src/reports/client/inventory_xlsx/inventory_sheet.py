"""Hoja de inventario completo."""

from openpyxl.styles import Alignment

from .styles import TITLE_FONT, SUB_FONT, _fill
from .helpers import _set_col_widths, _write_header_row, _write_data_row, _severity_fill, _state_fill


def build_inventory_sheet(wb, *, generated: str, total: int, resources: list):
    ws = wb.create_sheet("Inventario Completo")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [
        (1, 24), (2, 16), (3, 20), (4, 38), (5, 18),
        (6, 14), (7, 13), (8, 11), (9, 13), (10, 18), (11, 36), (12, 14), (13, 14),
    ])

    ws.merge_cells("A1:M1")
    ws["A1"] = "Inventario Completo de Recursos AWS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Generado: {generated}  ·  Total recursos activos: {total}"
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[2].height = 14

    cols = [
        "Cuenta AWS", "Servicio", "Tipo de Recurso", "ID de Recurso",
        "Región", "Estado", "Con Hallazgos", "Nº Hallazgos",
        "Sev. Máx.", "Ahorro Est. (USD/mes)", "Tags", "Detectado", "Última vez visto",
    ]
    _write_header_row(ws, 4, cols)
    ws.freeze_panes = "A5"

    for i, r in enumerate(resources):
        row   = 5 + i
        has_f = r.get("has_findings", False)
        state = r.get("state", "unknown")
        sev   = r.get("max_severity", "—")

        color_map = {}
        if has_f:
            color_map[7] = _fill("FEE2E2")
            color_map[9] = _severity_fill(sev)
        color_map[6] = _state_fill(state)

        _write_data_row(ws, row, [
            r.get("account_name", ""),
            r.get("service_name", ""),
            r.get("resource_type", ""),
            r.get("resource_id", ""),
            r.get("region", ""),
            state.capitalize(),
            "Sí" if has_f else "No",
            r.get("findings_count", 0),
            sev,
            f"${r.get('est_savings', 0):,.2f}",
            r.get("tags", "—"),
            r.get("detected_at", "—"),
            r.get("last_seen_at", "—"),
        ], alt=(i % 2 == 0), wrap_cols={11}, color_map=color_map, row_h=20)
