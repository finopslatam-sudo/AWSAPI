"""
Style constants for the inventory XLSX report.
Fills, fonts, borders, and low-level helpers.
"""

from openpyxl.styles import Font, PatternFill, Border, Side


# ─────────────────────────────────────────────────────────
# BORDER HELPERS
# ─────────────────────────────────────────────────────────

def _side() -> Side:
    return Side(style="thin", color="CBD5E1")


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border() -> Border:
    t = Side(style="medium", color="94A3B8")
    n = _side()
    return Border(left=t, right=t, top=t, bottom=n)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


# ─────────────────────────────────────────────────────────
# FILL CONSTANTS
# ─────────────────────────────────────────────────────────

DARK_FILL    = _fill("1E293B")
ACCENT_FILL  = _fill("2563EB")
ALT_FILL     = _fill("F8FAFC")
WHITE_FILL   = _fill("FFFFFF")
GREEN_FILL   = _fill("DCFCE7")
RED_FILL     = _fill("FEE2E2")
AMBER_FILL   = _fill("FEF3C7")
BLUE_FILL    = _fill("DBEAFE")
PURPLE_FILL  = _fill("F3E8FF")
SUMMARY_FILL = _fill("EFF6FF")

# ─────────────────────────────────────────────────────────
# FONT CONSTANTS
# ─────────────────────────────────────────────────────────

HDR_FONT     = Font(color="FFFFFF", bold=True, size=9)
LABEL_FONT   = Font(bold=True, color="0F172A", size=9)
VALUE_FONT   = Font(color="334155", size=9)
TITLE_FONT   = Font(bold=True, size=15, color="0F172A")
SUB_FONT     = Font(size=9, color="64748B")
KPI_FONT     = Font(bold=True, size=13, color="1D4ED8")
SECTION_FONT = Font(bold=True, size=11, color="1E293B")

# ─────────────────────────────────────────────────────────
# BORDER CONSTANT
# ─────────────────────────────────────────────────────────

THIN = _border()
