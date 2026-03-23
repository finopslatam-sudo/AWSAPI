"""
Worksheet builders for the inventory XLSX report.
Each function receives the workbook and extracted stats fields,
adds one sheet, and returns nothing.
"""

from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference

from .styles import (
    TITLE_FONT, SUB_FONT, SECTION_FONT, LABEL_FONT,
    ACCENT_FILL, BLUE_FILL, RED_FILL, GREEN_FILL,
    AMBER_FILL, PURPLE_FILL, SUMMARY_FILL, THIN,
    _fill,
)
from .helpers import (
    _set_col_widths, _write_header_row, _write_data_row,
    _kpi_block, _severity_fill, _state_fill,
)


# ─────────────────────────────────────────────────────────
# HOJA 1 — RESUMEN EJECUTIVO
# ─────────────────────────────────────────────────────────

def build_summary_sheet(ws, *, generated: str, plan: str, acc_count: int,
                        acc_label: str, users: int, total: int, with_f: int,
                        without_f: int, active_f_count: int, total_savings: float,
                        by_service: dict, by_state: dict, by_region: dict):

    ws.title = "Resumen Ejecutivo"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [
        (1, 4), (2, 28), (3, 22), (4, 4), (5, 28), (6, 22), (7, 4), (8, 28), (9, 22),
    ])

    # — título principal —
    ws.merge_cells("B1:I1")
    ws["B1"] = "Inventario de Recursos AWS — FinOpsLatam"
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B2:I2")
    ws["B2"] = (
        f"Generado: {generated}  ·  Plan: {plan}  ·  "
        f"Cuentas: {acc_count} ({acc_label})  ·  Usuarios: {users}"
    )
    ws["B2"].font = SUB_FONT
    ws["B2"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 16

    # — separador —
    for col in range(2, 10):
        ws.cell(row=3, column=col).fill = _fill("2563EB")
    ws.row_dimensions[3].height = 3

    # — bloque KPIs —
    total_s = max(total, 1)
    kpi_data = [
        (5, 2, "TOTAL RECURSOS",      str(total),           "activos en inventario",                               BLUE_FILL,          "1D4ED8"),
        (5, 5, "CON HALLAZGOS",       str(with_f),          f"{(with_f/total_s*100):.1f}% del inventario",         RED_FILL,           "DC2626"),
        (5, 8, "SIN HALLAZGOS",       str(without_f),       f"{(without_f/total_s*100):.1f}% sin alertas",         GREEN_FILL,         "15803D"),
        (9, 2, "HALLAZGOS ACTIVOS",   str(active_f_count),  "findings sin resolver",                               AMBER_FILL,         "B45309"),
        (9, 5, "AHORRO POTENCIAL",    f"USD ${total_savings:,.2f}", "estimado mensual",                            _fill("F0FDF4"),    "059669"),
        (9, 8, "SERVICIOS DISTINTOS", str(len(by_service)), "tipos de servicio AWS",                               PURPLE_FILL,        "6D28D9"),
    ]
    for r, c, lbl, val, note, fill, color in kpi_data:
        for dr in range(3):
            ws.merge_cells(
                start_row=r+dr, start_column=c,
                end_row=r+dr,   end_column=c+1,
            )
        _kpi_block(ws, r, c, lbl, val, note, fill, color)
        for dr in range(3):
            for dc in range(2):
                ws.cell(row=r+dr, column=c+dc).border = THIN

    # — distribución por estado —
    state_start = 13
    ws.merge_cells(f"B{state_start}:I{state_start}")
    ws[f"B{state_start}"] = "Distribución por Estado"
    ws[f"B{state_start}"].font = SECTION_FONT
    ws.row_dimensions[state_start].height = 20

    _write_header_row(ws, state_start + 1, ["Estado", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    for i, (state, count) in enumerate(by_state.items()):
        r = state_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws, r,
                        [state.capitalize(), count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0),
                        color_map={1: _state_fill(state)})

    # — distribución por servicio (top 12) —
    svc_start = state_start + 2 + len(by_state) + 2
    ws.merge_cells(f"B{svc_start}:I{svc_start}")
    ws[f"B{svc_start}"] = "Top Servicios AWS por Recursos"
    ws[f"B{svc_start}"].font = SECTION_FONT
    ws.row_dimensions[svc_start].height = 20

    _write_header_row(ws, svc_start + 1,
                      ["Servicio AWS", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    top_svcs = list(by_service.items())[:12]
    for i, (svc, count) in enumerate(top_svcs):
        r = svc_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws, r,
                        [svc, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))

    # — distribución por región (top 12) —
    reg_start = svc_start + 2 + len(top_svcs) + 2
    ws.merge_cells(f"B{reg_start}:I{reg_start}")
    ws[f"B{reg_start}"] = "Distribución por Región Geográfica"
    ws[f"B{reg_start}"].font = SECTION_FONT
    ws.row_dimensions[reg_start].height = 20

    _write_header_row(ws, reg_start + 1,
                      ["Región AWS", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    top_regs = list(by_region.items())[:12]
    for i, (reg, count) in enumerate(top_regs):
        r = reg_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws, r,
                        [reg, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))


# ─────────────────────────────────────────────────────────
# HOJA 2 — INVENTARIO COMPLETO
# ─────────────────────────────────────────────────────────

def build_inventory_sheet(wb, *, generated: str, total: int, resources: list):
    ws = wb.create_sheet("Inventario Completo")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [
        (1, 24), (2, 16), (3, 20), (4, 38), (5, 18),
        (6, 14), (7, 13), (8, 11), (9, 13), (10, 18), (11, 36), (12, 14), (13, 14),
    ])

    ws.merge_cells("A1:M1")
    ws["A1"] = "Inventario Completo de Recursos AWS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Generado: {generated}  ·  Total recursos activos: {total}"
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[2].height = 14

    COLS2 = [
        "Cuenta AWS", "Servicio", "Tipo de Recurso", "ID de Recurso",
        "Región", "Estado", "Con Hallazgos", "Nº Hallazgos",
        "Sev. Máx.", "Ahorro Est. (USD/mes)", "Tags", "Detectado", "Última vez visto",
    ]
    _write_header_row(ws, 4, COLS2)
    ws.freeze_panes = "A5"

    for i, r in enumerate(resources):
        row   = 5 + i
        has_f = r.get("has_findings", False)
        state = r.get("state", "unknown")
        sev   = r.get("max_severity", "—")

        color_map = {}
        if has_f:
            color_map[7] = _fill("FEE2E2")
            color_map[9] = _severity_fill(sev)
        color_map[6] = _state_fill(state)

        _write_data_row(ws, row, [
            r.get("account_name", ""),
            r.get("service_name", ""),
            r.get("resource_type", ""),
            r.get("resource_id", ""),
            r.get("region", ""),
            state.capitalize(),
            "Sí" if has_f else "No",
            r.get("findings_count", 0),
            sev,
            f"${r.get('est_savings', 0):,.2f}",
            r.get("tags", "—"),
            r.get("detected_at", "—"),
            r.get("last_seen_at", "—"),
        ], alt=(i % 2 == 0), wrap_cols={11}, color_map=color_map, row_h=20)


# ─────────────────────────────────────────────────────────
# HOJA 3 — POR SERVICIO AWS
# ─────────────────────────────────────────────────────────

def build_service_sheet(wb, *, generated: str, total: int,
                        by_service: dict, resources: list):
    ws = wb.create_sheet("Por Servicio AWS")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [(1, 32), (2, 14), (3, 14), (4, 28)])

    ws.merge_cells("A1:D1")
    ws["A1"] = "Inventario por Servicio AWS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generado: {generated}"
    ws["A2"].font = SUB_FONT

    _write_header_row(ws, 4, ["Servicio AWS", "Recursos", "% del Total", "Distribución"])

    total_s = max(total, 1)
    svc_chart_data = []
    for i, (svc, count) in enumerate(by_service.items()):
        row = 5 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 3)) + "░" * (33 - max(1, int(pct / 3)))
        _write_data_row(ws, row,
                        [svc, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))
        svc_chart_data.append((svc, count))

    # fila total
    total_row = 5 + len(by_service)
    _write_data_row(ws, total_row,
                    ["TOTAL", total, "100%", ""],
                    color_map={1: SUMMARY_FILL, 2: SUMMARY_FILL, 3: SUMMARY_FILL, 4: SUMMARY_FILL})
    for c in range(1, 5):
        ws.cell(total_row, c).font = LABEL_FONT

    # gráfico de barras
    if svc_chart_data:
        chart = BarChart()
        chart.type     = "bar"
        chart.title    = "Recursos por Servicio AWS"
        chart.y_axis.title = "Servicio"
        chart.x_axis.title = "Recursos"
        chart.style    = 10
        chart.width    = 18
        chart.height   = max(8, len(svc_chart_data) * 0.65)
        chart.grouping = "clustered"

        data_ref = Reference(ws, min_col=2, min_row=4,
                             max_col=2, max_row=4 + len(svc_chart_data))
        cats_ref = Reference(ws, min_col=1, min_row=5,
                             max_row=4 + len(svc_chart_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "2563EB"
        ws.add_chart(chart, "F4")


# ─────────────────────────────────────────────────────────
# HOJA 4 — POR REGIÓN
# ─────────────────────────────────────────────────────────

def build_region_sheet(wb, *, generated: str, total: int, by_region: dict):
    ws = wb.create_sheet("Por Región")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [(1, 28), (2, 14), (3, 14), (4, 28)])

    ws.merge_cells("A1:D1")
    ws["A1"] = "Inventario por Región Geográfica"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generado: {generated}"
    ws["A2"].font = SUB_FONT

    _write_header_row(ws, 4, ["Región AWS", "Recursos", "% del Total", "Distribución"])

    total_s = max(total, 1)
    reg_chart_data = []
    for i, (reg, count) in enumerate(by_region.items()):
        row = 5 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 3)) + "░" * (33 - max(1, int(pct / 3)))
        _write_data_row(ws, row,
                        [reg, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))
        reg_chart_data.append((reg, count))

    total_row4 = 5 + len(by_region)
    _write_data_row(ws, total_row4,
                    ["TOTAL", total, "100%", ""],
                    color_map={1: SUMMARY_FILL, 2: SUMMARY_FILL, 3: SUMMARY_FILL, 4: SUMMARY_FILL})
    for c in range(1, 5):
        ws.cell(total_row4, c).font = LABEL_FONT

    if reg_chart_data:
        chart4 = BarChart()
        chart4.type     = "bar"
        chart4.title    = "Recursos por Región AWS"
        chart4.y_axis.title = "Región"
        chart4.x_axis.title = "Recursos"
        chart4.style    = 10
        chart4.width    = 18
        chart4.height   = max(8, len(reg_chart_data) * 0.8)
        chart4.grouping = "clustered"

        data_ref4 = Reference(ws, min_col=2, min_row=4,
                              max_col=2, max_row=4 + len(reg_chart_data))
        cats_ref4 = Reference(ws, min_col=1, min_row=5,
                              max_row=4 + len(reg_chart_data))
        chart4.add_data(data_ref4, titles_from_data=True)
        chart4.set_categories(cats_ref4)
        chart4.series[0].graphicalProperties.solidFill = "0EA5E9"
        ws.add_chart(chart4, "F4")


# ─────────────────────────────────────────────────────────
# HOJA 5 — CON HALLAZGOS
# ─────────────────────────────────────────────────────────

def build_findings_sheet(wb, *, generated: str, total_savings: float, resources: list):
    ws = wb.create_sheet("Con Hallazgos")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [
        (1, 22), (2, 15), (3, 20), (4, 36),
        (5, 16), (6, 13), (7, 11), (8, 16),
        (9, 22), (10, 18), (11, 48),
    ])

    ws.merge_cells("A1:K1")
    ws["A1"] = "Recursos con Hallazgos Activos"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    flagged = [r for r in resources if r.get("has_findings")]

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        f"Generado: {generated}  ·  "
        f"{len(flagged)} recursos con hallazgos activos  ·  "
        f"Ahorro potencial total: USD ${total_savings:,.2f}/mes"
    )
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[2].height = 14

    COLS5 = [
        "Cuenta AWS", "Servicio", "Tipo", "ID de Recurso",
        "Región", "Estado", "Nº Hallazgos", "Ahorro Est. (USD/mes)",
        "Tipo de Finding", "Severidad", "Descripción",
    ]
    _write_header_row(ws, 4, COLS5, fill=_fill("DC2626"))
    ws.freeze_panes = "A5"

    data_row = 5
    for r in flagged:
        findings = r.get("findings", [])
        if not findings:
            continue
        for j, f in enumerate(findings):
            sev = f.get("severity", "—").upper()
            _write_data_row(ws, data_row, [
                r.get("account_name", "")           if j == 0 else "",
                r.get("service_name", "")           if j == 0 else "",
                r.get("resource_type", "")          if j == 0 else "",
                r.get("resource_id", "")            if j == 0 else "",
                r.get("region", "")                 if j == 0 else "",
                r.get("state", "").capitalize()     if j == 0 else "",
                r.get("findings_count", 0)          if j == 0 else "",
                f"${r.get('est_savings', 0):,.2f}"  if j == 0 else "",
                f.get("type", ""),
                sev,
                f.get("message", ""),
            ], alt=(data_row % 2 == 0),
               wrap_cols={11},
               row_h=22,
               color_map={10: _severity_fill(sev)})
            data_row += 1
