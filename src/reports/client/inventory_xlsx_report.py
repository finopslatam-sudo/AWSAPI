"""
INVENTORY XLSX REPORT — ENTERPRISE
====================================
5 hojas:
  1. Resumen Ejecutivo   — KPIs + distribuciones
  2. Inventario Completo — todos los recursos activos
  3. Por Servicio AWS    — agrupado por servicio con %
  4. Por Región          — agrupado por región con %
  5. Con Hallazgos       — recursos con findings activos + detalle
"""

from io import BytesIO
from datetime import datetime
import pytz

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel


# ─────────────────────────────────────────────────────────
# ESTILOS BASE
# ─────────────────────────────────────────────────────────

def _side():
    return Side(style="thin", color="CBD5E1")


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border():
    t = Side(style="medium", color="94A3B8")
    n = _side()
    return Border(left=t, right=t, top=t, bottom=n)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


DARK_FILL    = _fill("1E293B")
ACCENT_FILL  = _fill("2563EB")
ALT_FILL     = _fill("F8FAFC")
WHITE_FILL   = _fill("FFFFFF")
GREEN_FILL   = _fill("DCFCE7")
RED_FILL     = _fill("FEE2E2")
AMBER_FILL   = _fill("FEF3C7")
BLUE_FILL    = _fill("DBEAFE")
PURPLE_FILL  = _fill("F3E8FF")
SUMMARY_FILL = _fill("EFF6FF")

HDR_FONT   = Font(color="FFFFFF", bold=True, size=9)
LABEL_FONT = Font(bold=True, color="0F172A", size=9)
VALUE_FONT = Font(color="334155", size=9)
TITLE_FONT = Font(bold=True, size=15, color="0F172A")
SUB_FONT   = Font(size=9, color="64748B")
KPI_FONT   = Font(bold=True, size=13, color="1D4ED8")
SECTION_FONT = Font(bold=True, size=11, color="1E293B")

THIN = _border()


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────

def _set_col_widths(ws, widths: list[tuple[int, float]]):
    for col_idx, w in widths:
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _write_header_row(ws, row: int, labels: list[str], fill=None):
    f = fill or DARK_FILL
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HDR_FONT
        cell.fill = f
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20


def _write_data_row(ws, row: int, values: list, alt: bool = False,
                    wrap_cols: set | None = None, row_h: int = 18,
                    color_map: dict | None = None):
    bg = ALT_FILL if alt else WHITE_FILL
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = VALUE_FONT
        cell.fill = color_map.get(c, bg) if color_map else bg
        cell.border = THIN
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(wrap_cols is not None and c in wrap_cols),
        )
    ws.row_dimensions[row].height = row_h


def _kpi_block(ws, start_row: int, start_col: int, label: str, value: str,
               note: str, fill: PatternFill, val_color: str = "1D4ED8"):
    """Escribe un bloque KPI de 3 filas × 1 columna."""
    r = start_row
    c = start_col

    lbl_cell = ws.cell(row=r,   column=c, value=label)
    val_cell = ws.cell(row=r+1, column=c, value=value)
    note_cell= ws.cell(row=r+2, column=c, value=note)

    lbl_cell.font  = Font(size=8, color="64748B")
    val_cell.font  = Font(bold=True, size=14, color=val_color)
    note_cell.font = Font(size=7, color="94A3B8")

    for cell in (lbl_cell, val_cell, note_cell):
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[r].height   = 14
    ws.row_dimensions[r+1].height = 22
    ws.row_dimensions[r+2].height = 13


def _severity_fill(severity: str) -> PatternFill:
    s = (severity or "").upper()
    if s == "CRITICAL":
        return _fill("FFE4E6")
    if s == "HIGH":
        return _fill("FEE2E2")
    if s == "MEDIUM":
        return _fill("FEF3C7")
    if s == "LOW":
        return _fill("DCFCE7")
    return WHITE_FILL


def _state_fill(state: str) -> PatternFill:
    s = (state or "").lower()
    if s in ("running", "available", "active"):
        return _fill("DCFCE7")
    if s in ("stopped", "stopping"):
        return _fill("FEE2E2")
    if s in ("idle", "inactive"):
        return _fill("FEF3C7")
    return ALT_FILL


# ─────────────────────────────────────────────────────────
# BUILD
# ─────────────────────────────────────────────────────────

def build_inventory_xlsx(stats: dict) -> bytes:

    chile_tz  = pytz.timezone("America/Santiago")
    generated = datetime.now(chile_tz).strftime("%d/%m/%Y %H:%M CLT")

    resources      = stats.get("resources", [])
    by_service     = stats.get("by_service", {})
    by_region      = stats.get("by_region", {})
    by_state       = stats.get("by_state", {})
    total          = stats.get("total", 0)
    with_f         = stats.get("with_findings", 0)
    without_f      = stats.get("without_findings", 0)
    total_savings  = stats.get("total_savings", 0.0)
    active_f_count = stats.get("active_findings_count", 0)
    plan           = stats.get("plan", "—")
    users          = stats.get("user_count", 0)
    acc_count      = stats.get("account_count", 0)
    acc_label      = stats.get("account_label", "Todas las cuentas")

    wb = Workbook()

    # ═══════════════════════════════════════════════════════
    # HOJA 1 — RESUMEN EJECUTIVO
    # ═══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_view.showGridLines = False
    _set_col_widths(ws1, [
        (1, 4), (2, 28), (3, 22), (4, 4), (5, 28), (6, 22), (7, 4), (8, 28), (9, 22),
    ])

    # — título principal —
    ws1.merge_cells("B1:I1")
    ws1["B1"] = "Inventario de Recursos AWS — FinOpsLatam"
    ws1["B1"].font = TITLE_FONT
    ws1["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("B2:I2")
    ws1["B2"] = f"Generado: {generated}  ·  Plan: {plan}  ·  Cuentas: {acc_count} ({acc_label})  ·  Usuarios: {users}"
    ws1["B2"].font = SUB_FONT
    ws1["B2"].alignment = Alignment(horizontal="left")
    ws1.row_dimensions[2].height = 16

    # — separador —
    for col in range(2, 10):
        c = ws1.cell(row=3, column=col)
        c.fill = _fill("2563EB")
    ws1.row_dimensions[3].height = 3

    # — bloque KPIs (fila 5, 3 columnas) —
    kpi_data = [
        (5, 2, "TOTAL RECURSOS",        str(total),           "activos en inventario",     BLUE_FILL,   "1D4ED8"),
        (5, 5, "CON HALLAZGOS",         str(with_f),          f"{(with_f/max(total,1)*100):.1f}% del inventario", RED_FILL,    "DC2626"),
        (5, 8, "SIN HALLAZGOS",         str(without_f),       f"{(without_f/max(total,1)*100):.1f}% sin alertas", GREEN_FILL,  "15803D"),
        (9, 2, "HALLAZGOS ACTIVOS",     str(active_f_count),  "findings sin resolver",     AMBER_FILL,  "B45309"),
        (9, 5, "AHORRO POTENCIAL",      f"USD ${total_savings:,.2f}", "estimado mensual",  _fill("F0FDF4"), "059669"),
        (9, 8, "SERVICIOS DISTINTOS",   str(len(by_service)), "tipos de servicio AWS",     PURPLE_FILL, "6D28D9"),
    ]
    for r, c, lbl, val, note, fill, color in kpi_data:
        ws1.merge_cells(
            start_row=r,   start_column=c,
            end_row=r,     end_column=c+1,
        )
        ws1.merge_cells(
            start_row=r+1, start_column=c,
            end_row=r+1,   end_column=c+1,
        )
        ws1.merge_cells(
            start_row=r+2, start_column=c,
            end_row=r+2,   end_column=c+1,
        )
        _kpi_block(ws1, r, c, lbl, val, note, fill, color)
        for dr in range(3):
            for dc in range(2):
                ws1.cell(row=r+dr, column=c+dc).border = THIN

    # — distribución por estado —
    state_start = 13
    ws1.merge_cells(f"B{state_start}:I{state_start}")
    ws1[f"B{state_start}"] = "Distribución por Estado"
    ws1[f"B{state_start}"].font = SECTION_FONT
    ws1.row_dimensions[state_start].height = 20

    _write_header_row(ws1, state_start + 1, ["Estado", "Recursos", "% del Total", "Indicador"], fill=ACCENT_FILL)
    total_s = max(total, 1)
    for i, (state, count) in enumerate(by_state.items()):
        r = state_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws1, r,
                        [state.capitalize(), count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0),
                        color_map={1: _state_fill(state)})

    # — distribución por servicio (top 10) —
    svc_start = state_start + 2 + len(by_state) + 2
    ws1.merge_cells(f"B{svc_start}:I{svc_start}")
    ws1[f"B{svc_start}"] = "Top Servicios AWS por Recursos"
    ws1[f"B{svc_start}"].font = SECTION_FONT
    ws1.row_dimensions[svc_start].height = 20

    _write_header_row(ws1, svc_start + 1,
                      ["Servicio AWS", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    top_svcs = list(by_service.items())[:12]
    for i, (svc, count) in enumerate(top_svcs):
        r = svc_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws1, r,
                        [svc, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))

    # — distribución por región (top 10) —
    reg_start = svc_start + 2 + len(top_svcs) + 2
    ws1.merge_cells(f"B{reg_start}:I{reg_start}")
    ws1[f"B{reg_start}"] = "Distribución por Región Geográfica"
    ws1[f"B{reg_start}"].font = SECTION_FONT
    ws1.row_dimensions[reg_start].height = 20

    _write_header_row(ws1, reg_start + 1,
                      ["Región AWS", "Recursos", "% del Total", "Indicador"],
                      fill=ACCENT_FILL)
    top_regs = list(by_region.items())[:12]
    for i, (reg, count) in enumerate(top_regs):
        r = reg_start + 2 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 5)) + "░" * (20 - max(1, int(pct / 5)))
        _write_data_row(ws1, r,
                        [reg, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))

    # ═══════════════════════════════════════════════════════
    # HOJA 2 — INVENTARIO COMPLETO
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Inventario Completo")
    ws2.sheet_view.showGridLines = False
    _set_col_widths(ws2, [
        (1, 24), (2, 16), (3, 20), (4, 38), (5, 18),
        (6, 14), (7, 13), (8, 11), (9, 13), (10, 18), (11, 36), (12, 14), (13, 14),
    ])

    ws2.merge_cells("A1:M1")
    ws2["A1"] = "Inventario Completo de Recursos AWS"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:M2")
    ws2["A2"] = f"Generado: {generated}  ·  Total recursos activos: {total}"
    ws2["A2"].font = SUB_FONT
    ws2.row_dimensions[2].height = 14

    COLS2 = [
        "Cuenta AWS", "Servicio", "Tipo de Recurso", "ID de Recurso",
        "Región", "Estado", "Con Hallazgos", "Nº Hallazgos",
        "Sev. Máx.", "Ahorro Est. (USD/mes)", "Tags", "Detectado", "Última vez visto",
    ]
    _write_header_row(ws2, 4, COLS2)
    ws2.freeze_panes = "A5"

    for i, r in enumerate(resources):
        row = 5 + i
        has_f  = r.get("has_findings", False)
        state  = r.get("state", "unknown")
        sev    = r.get("max_severity", "—")

        color_map = {}
        if has_f:
            color_map[7] = _fill("FEE2E2")   # col "Con Hallazgos"
            color_map[9] = _severity_fill(sev)
        color_map[6] = _state_fill(state)

        _write_data_row(ws2, row, [
            r.get("account_name", ""),
            r.get("service_name", ""),
            r.get("resource_type", ""),
            r.get("resource_id", ""),
            r.get("region", ""),
            state.capitalize(),
            "Sí" if has_f else "No",
            r.get("findings_count", 0),
            sev,
            f"${r.get('est_savings', 0):,.2f}",
            r.get("tags", "—"),
            r.get("detected_at", "—"),
            r.get("last_seen_at", "—"),
        ], alt=(i % 2 == 0), wrap_cols={11}, color_map=color_map, row_h=20)

    # ═══════════════════════════════════════════════════════
    # HOJA 3 — POR SERVICIO AWS
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Por Servicio AWS")
    ws3.sheet_view.showGridLines = False
    _set_col_widths(ws3, [(1, 32), (2, 14), (3, 14), (4, 28)])

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Inventario por Servicio AWS"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 28

    ws3.merge_cells("A2:D2")
    ws3["A2"] = f"Generado: {generated}"
    ws3["A2"].font = SUB_FONT

    _write_header_row(ws3, 4, ["Servicio AWS", "Recursos", "% del Total", "Distribución"])

    # cuenta recursos con findings por servicio
    findings_by_svc: dict[str, int] = {}
    for r in resources:
        if r.get("has_findings"):
            findings_by_svc[r["service_name"]] = findings_by_svc.get(r["service_name"], 0) + 1

    total_s = max(total, 1)
    svc_chart_data = []
    for i, (svc, count) in enumerate(by_service.items()):
        row  = 5 + i
        pct  = count / total_s * 100
        bar  = "█" * max(1, int(pct / 3)) + "░" * (33 - max(1, int(pct / 3)))
        _write_data_row(ws3, row,
                        [svc, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))
        svc_chart_data.append((svc, count))

    # fila total
    total_row = 5 + len(by_service)
    _write_data_row(ws3, total_row,
                    ["TOTAL", total, "100%", ""],
                    color_map={1: SUMMARY_FILL, 2: SUMMARY_FILL, 3: SUMMARY_FILL, 4: SUMMARY_FILL})
    for c in range(1, 5):
        ws3.cell(total_row, c).font = LABEL_FONT

    # gráfico de barras (si hay datos)
    if svc_chart_data:
        chart = BarChart()
        chart.type   = "bar"
        chart.title  = "Recursos por Servicio AWS"
        chart.y_axis.title = "Servicio"
        chart.x_axis.title = "Recursos"
        chart.style  = 10
        chart.width  = 18
        chart.height = max(8, len(svc_chart_data) * 0.65)
        chart.grouping = "clustered"

        data_ref = Reference(ws3,
                             min_col=2, min_row=4,
                             max_col=2, max_row=4 + len(svc_chart_data))
        cats_ref = Reference(ws3,
                             min_col=1, min_row=5,
                             max_row=4 + len(svc_chart_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "2563EB"
        ws3.add_chart(chart, f"F4")

    # ═══════════════════════════════════════════════════════
    # HOJA 4 — POR REGIÓN
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Por Región")
    ws4.sheet_view.showGridLines = False
    _set_col_widths(ws4, [(1, 28), (2, 14), (3, 14), (4, 28)])

    ws4.merge_cells("A1:D1")
    ws4["A1"] = "Inventario por Región Geográfica"
    ws4["A1"].font = TITLE_FONT
    ws4["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells("A2:D2")
    ws4["A2"] = f"Generado: {generated}"
    ws4["A2"].font = SUB_FONT

    _write_header_row(ws4, 4, ["Región AWS", "Recursos", "% del Total", "Distribución"])

    reg_chart_data = []
    for i, (reg, count) in enumerate(by_region.items()):
        row = 5 + i
        pct = count / total_s * 100
        bar = "█" * max(1, int(pct / 3)) + "░" * (33 - max(1, int(pct / 3)))
        _write_data_row(ws4, row,
                        [reg, count, f"{pct:.1f}%", bar],
                        alt=(i % 2 == 0))
        reg_chart_data.append((reg, count))

    total_row4 = 5 + len(by_region)
    _write_data_row(ws4, total_row4,
                    ["TOTAL", total, "100%", ""],
                    color_map={1: SUMMARY_FILL, 2: SUMMARY_FILL, 3: SUMMARY_FILL, 4: SUMMARY_FILL})
    for c in range(1, 5):
        ws4.cell(total_row4, c).font = LABEL_FONT

    if reg_chart_data:
        chart4 = BarChart()
        chart4.type   = "bar"
        chart4.title  = "Recursos por Región AWS"
        chart4.y_axis.title = "Región"
        chart4.x_axis.title = "Recursos"
        chart4.style  = 10
        chart4.width  = 18
        chart4.height = max(8, len(reg_chart_data) * 0.8)
        chart4.grouping = "clustered"

        data_ref4 = Reference(ws4,
                              min_col=2, min_row=4,
                              max_col=2, max_row=4 + len(reg_chart_data))
        cats_ref4 = Reference(ws4,
                              min_col=1, min_row=5,
                              max_row=4 + len(reg_chart_data))
        chart4.add_data(data_ref4, titles_from_data=True)
        chart4.set_categories(cats_ref4)
        chart4.series[0].graphicalProperties.solidFill = "0EA5E9"
        ws4.add_chart(chart4, "F4")

    # ═══════════════════════════════════════════════════════
    # HOJA 5 — CON HALLAZGOS
    # ═══════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Con Hallazgos")
    ws5.sheet_view.showGridLines = False
    _set_col_widths(ws5, [
        (1, 22), (2, 15), (3, 20), (4, 36),
        (5, 16), (6, 13), (7, 11), (8, 16),
        (9, 22), (10, 18), (11, 48),
    ])

    ws5.merge_cells("A1:K1")
    ws5["A1"] = "Recursos con Hallazgos Activos"
    ws5["A1"].font = TITLE_FONT
    ws5["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws5.row_dimensions[1].height = 28

    flagged = [r for r in resources if r.get("has_findings")]

    ws5.merge_cells("A2:K2")
    ws5["A2"] = (
        f"Generado: {generated}  ·  "
        f"{len(flagged)} recursos con hallazgos activos  ·  "
        f"Ahorro potencial total: USD ${total_savings:,.2f}/mes"
    )
    ws5["A2"].font = SUB_FONT
    ws5.row_dimensions[2].height = 14

    COLS5 = [
        "Cuenta AWS", "Servicio", "Tipo", "ID de Recurso",
        "Región", "Estado", "Nº Hallazgos", "Ahorro Est. (USD/mes)",
        "Tipo de Finding", "Severidad", "Descripción",
    ]
    _write_header_row(ws5, 4, COLS5, fill=_fill("DC2626"))
    ws5.freeze_panes = "A5"

    data_row = 5
    for r in flagged:
        findings = r.get("findings", [])
        if not findings:
            continue
        for j, f in enumerate(findings):
            sev = f.get("severity", "—").upper()
            _write_data_row(ws5, data_row, [
                r.get("account_name", "")    if j == 0 else "",
                r.get("service_name", "")    if j == 0 else "",
                r.get("resource_type", "")   if j == 0 else "",
                r.get("resource_id", "")     if j == 0 else "",
                r.get("region", "")          if j == 0 else "",
                r.get("state", "").capitalize() if j == 0 else "",
                r.get("findings_count", 0)   if j == 0 else "",
                f"${r.get('est_savings', 0):,.2f}" if j == 0 else "",
                f.get("type", ""),
                sev,
                f.get("message", ""),
            ], alt=(data_row % 2 == 0),
               wrap_cols={11},
               row_h=22,
               color_map={10: _severity_fill(sev)})
            data_row += 1

    # ═══════════════════════════════════════════════════════
    # EXPORT
    # ═══════════════════════════════════════════════════════
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
